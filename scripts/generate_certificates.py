"""
SpaceBorn Certificate Generator
--------------------------------
Reads rows from certificates_data.xlsx and generates one certificate
image (PNG) per row, styled after the SpaceBorn "Certificate of
Completion" template.

USAGE:
    python3 generate_certificates.py

    - Add/update rows in certificates_data.xlsx (keep the header row).
    - Run this script. New/updated certificates are written to the
      "output" folder as "<Name>.png".
    - Already-generated certificates are skipped unless you delete
      the PNG or force regeneration with --force.

Requires: pip install openpyxl pillow qrcode --break-system-packages
"""

import argparse
import csv
import json
import os
import re
import textwrap
import urllib.error
import urllib.request
import uuid
from datetime import datetime

import openpyxl
import qrcode
from PIL import Image, ImageDraw, ImageFont

# ----------------------------- CONFIG -----------------------------

EXCEL_FILE = "certificates_data.xlsx"
OUTPUT_DIR = "output"

# Path to the SpaceBorn seal watermark (the faint circular badge that sits
# bottom-left on the certificate). Place your seal PNG here.
SEAL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "seal.png")

# How see-through the seal watermark should be (0 = invisible, 255 = fully opaque)
SEAL_OPACITY = 40

# Diameter (in original 754px-wide coordinate space) of the seal watermark.
# It's intentionally oversized so it bleeds off the bottom/left edge, matching
# the reference certificate.
SEAL_DIAMETER = 460

# Cache so we don't reload/resize the seal image once per certificate
_seal_cache = {}

# --------------------------- QR CODE CONFIG ---------------------------

# Base verification URL. Each certificate's QR encodes (path style):
#     QR_BASE_URL/<certificate_id>
# e.g. "https://spaceborn.in/verify/8f3a1c9d2b41"
# >>> Replace this with your real verification domain once it's live. <<<
QR_BASE_URL = "https://spaceborn.in/verify"

# Size (in original 754px-wide coordinate space) of the QR code square,
# top-right corner.
QR_SIZE = 95

# Distance from the top and right edges (same coordinate space as above)
QR_MARGIN_TOP = 24
QR_MARGIN_RIGHT = 50

# Name of the Excel column that stores each row's permanent unique
# certificate ID. If this column/header doesn't exist yet, it's created
# automatically. If a row doesn't have an ID yet, one is generated once
# and written back into the sheet so it never changes on re-runs.
CERT_ID_HEADER = "Certificate ID"

# CSV written alongside the certificates every run — one row per
# certificate with its ID, all details, and the exact QR URL. Hand this
# to the backend team to seed/sync the DB table.
LOG_FILE = "certificates_log.csv"

# --------------------------- WEBSITE SYNC CONFIG ---------------------------

# After generating certificates, every row is pushed to the Spaceborn
# website so each QR code's /verify/<id> page shows the person's details.
# Set these two environment variables before running:
#
#   SPACEBORN_API_URL   e.g. "https://spaceborn.in"  (no trailing slash)
#   SPACEBORN_API_KEY   the same secret set as CERT_API_KEY on the server
#
# If SPACEBORN_API_KEY is not set, sync is skipped (certificates still
# generate normally). Use --no-sync to skip explicitly.
API_BASE_URL = os.environ.get("SPACEBORN_API_URL", "https://spaceborn.in").rstrip("/")
API_KEY = os.environ.get("SPACEBORN_API_KEY", "")

# Date format used in the Excel sheet (e.g. 22/02/2026)
DATE_FORMAT = "%d/%m/%Y"


def compute_experience(start_date, end_date):
    """Human-readable experience duration from the start/end dates,
    e.g. '3 Months' or '1 Year 2 Months'. Returns '' if dates don't parse."""
    try:
        start = datetime.strptime(start_date.strip(), DATE_FORMAT)
        end = datetime.strptime(end_date.strip(), DATE_FORMAT)
    except (ValueError, AttributeError):
        return ""
    if end < start:
        return ""

    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1

    if months <= 0:
        days = (end - start).days
        return f"{days} Day{'s' if days != 1 else ''}"

    years, rem_months = divmod(months, 12)
    parts = []
    if years:
        parts.append(f"{years} Year{'s' if years != 1 else ''}")
    if rem_months:
        parts.append(f"{rem_months} Month{'s' if rem_months != 1 else ''}")
    return " ".join(parts)


def sync_to_website(log_rows):
    """POST all certificate rows to the website's /api/certificates endpoint
    so each /verify/<id> page (linked from the QR code) shows live details."""
    if not API_KEY:
        print("Sync skipped: SPACEBORN_API_KEY not set. "
              "Set SPACEBORN_API_URL and SPACEBORN_API_KEY to enable website sync.")
        return

    url = f"{API_BASE_URL}/api/certificates"
    payload = {
        "certificates": [
            {
                "certificate_id": r["certificate_id"],
                "name": r["name"],
                "designation": r["designation"],
                "department": r["department"],
                "start_date": r["start_date"],
                "end_date": r["end_date"],
                "experience": r["experience"],
                "description": r["description"],
                "cert_date": r["cert_date"],
            }
            for r in log_rows
        ]
    }
    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "x-api-key": API_KEY},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            body = json.loads(response.read().decode("utf-8"))
        print(f"Synced {body.get('upserted', len(log_rows))} certificate(s) to {url} "
              f"({body.get('total', '?')} total on server).")
    except urllib.error.HTTPError as e:
        try:
            detail = json.loads(e.read().decode("utf-8")).get("error", "")
        except Exception:
            detail = ""
        print(f"WARNING: website sync failed ({e.code} {e.reason}). {detail}")
        print("Certificates were still generated locally; re-run to retry the sync.")
    except (urllib.error.URLError, TimeoutError, OSError) as e:
        print(f"WARNING: could not reach {url} ({e}).")
        print("Certificates were still generated locally; re-run to retry the sync.")


def slugify(text):
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def new_cert_id():
    # 12 hex chars from a uuid4 — short enough for a clean URL/QR,
    # long enough (2^48 space) to be collision-free at any real volume.
    return uuid.uuid4().hex[:12]

# Static company info (edit these once, applies to every certificate)
COMPANY_NAME = "SPACE BORN"
COMPANY_TAGLINE = "BEYOND AUTONOMOUS"
HEADER_LEFT = "Applied via SPICe+ | MCA. Govt of India | SRN: AC1678584"
SIGNER_NAME = "Adarsh Kumar"
SIGNER_TITLE = "Founder & CEO"
SIGNER_CONTACT = "Contact : adarshkumar@spaceborn.in"
ORG_LABEL = "From Spaceborn"

# Canvas size (rendered larger than the original 754x529 for print quality,
# then scaled proportionally)
W, H = 1600, 1123
SCALE = W / 754

# Candidate font file paths for each role, in priority order.
# Covers Windows (C:\Windows\Fonts), macOS, and Linux (DejaVu/Liberation).
FONT_CANDIDATES = {
    "sans": [
        "C:/Windows/Fonts/arial.ttf", "/Library/Fonts/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ],
    "sans_bold": [
        "C:/Windows/Fonts/arialbd.ttf", "/Library/Fonts/Arial Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ],
    "serif_bold": [
        "C:/Windows/Fonts/georgiab.ttf", "C:/Windows/Fonts/timesbd.ttf",
        "/Library/Fonts/Georgia Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf",
    ],
    "serif_italic": [
        "C:/Windows/Fonts/georgiai.ttf", "C:/Windows/Fonts/timesi.ttf",
        "/Library/Fonts/Georgia Italic.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSerif-Italic.ttf",
    ],
    "serif_bold_italic": [
        "C:/Windows/Fonts/georgiaz.ttf", "C:/Windows/Fonts/timesbi.ttf",
        "/Library/Fonts/Georgia Bold Italic.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSerif-BoldItalic.ttf",
    ],
}


def font(role, size):
    for path in FONT_CANDIDATES[role]:
        if os.path.exists(path):
            return ImageFont.truetype(path, int(size * SCALE))
    # last resort: PIL's bundled default (won't scale nicely, but won't crash)
    print(f"Warning: no font file found for role '{role}', using PIL default.")
    return ImageFont.load_default()


F_LOGO = font("sans_bold", 46)
F_TAGLINE = font("sans_bold", 16)
F_TITLE = font("serif_bold", 38)
F_SMALL = font("sans", 13)
F_PRESENTED = font("serif_italic", 16)
F_NAME = font("serif_bold_italic", 34)
F_BODY = font("sans", 14)
F_BODY_BOLD = font("sans_bold", 14)
F_DATE = font("sans_bold", 14)
F_SIGN = font("serif_italic", 20)
F_SIGNER = font("sans_bold", 14)
F_SIGNER_SMALL = font("sans", 12)

NAVY_DARK = (5, 15, 30)
NAVY_LIGHT = (20, 45, 75)
WHITE = (255, 255, 255)
LIGHT_GREY = (200, 210, 220)
ACCENT = (120, 160, 200)


def x(v):
    return int(v * SCALE)


def load_seal():
    """Load the seal PNG, resize to SEAL_DIAMETER, and fade its alpha to
    SEAL_OPACITY so it sits quietly behind the certificate text."""
    if "seal" in _seal_cache:
        return _seal_cache["seal"]

    if not os.path.exists(SEAL_PATH):
        print(f"Warning: seal image not found at '{SEAL_PATH}'. Skipping watermark.")
        _seal_cache["seal"] = None
        return None

    seal = Image.open(SEAL_PATH).convert("RGBA")
    size = x(SEAL_DIAMETER)
    seal = seal.resize((size, size), Image.LANCZOS)

    # scale down existing alpha channel to the target watermark opacity
    r, g, b, a = seal.split()
    a = a.point(lambda p: int(p * (SEAL_OPACITY / 255)))
    seal = Image.merge("RGBA", (r, g, b, a))

    _seal_cache["seal"] = seal
    return seal


def make_qr(data, size):
    """Generate a QR code PNG (as a PIL Image) encoding `data`, resized to
    a `size` x `size` square. The qrcode library already renders on a white
    background, so this drops in cleanly as a white box like the reference
    certificate."""
    qr = qrcode.QRCode(border=1, box_size=10)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    img = img.resize((size, size), Image.LANCZOS)
    return img


def draw_background(draw, img):
    # vertical gradient
    for y in range(H):
        t = y / H
        r = int(NAVY_DARK[0] + (NAVY_LIGHT[0] - NAVY_DARK[0]) * t)
        g = int(NAVY_DARK[1] + (NAVY_LIGHT[1] - NAVY_DARK[1]) * t)
        b = int(NAVY_DARK[2] + (NAVY_LIGHT[2] - NAVY_DARK[2]) * t)
        draw.line([(0, y), (W, y)], fill=(r, g, b))

    # faint seal watermark, bottom-left (intentionally bleeds off the edges)
    seal = load_seal()
    if seal is not None:
        cx, cy = x(60), H - x(20)
        paste_x = cx - seal.width // 2
        paste_y = cy - seal.height // 2
        img.paste(seal, (paste_x, paste_y), seal)

    # decorative circuit lines, top-right
    ox, oy = W - x(30), x(30)
    for i in range(6):
        end_x = ox - x(10) - i * x(28)
        end_y = oy + i * x(22)
        draw.line([(ox, oy), (end_x, oy), (end_x, end_y)], fill=NAVY_LIGHT, width=2)
        draw.ellipse([end_x - 4, end_y - 4, end_x + 4, end_y + 4], fill=NAVY_LIGHT)

    # border
    draw.rectangle([x(4), x(4), W - x(4), H - x(4)], outline=(40, 70, 100), width=2)


def wrap_text(draw, text, fnt, max_width):
    words = text.split()
    lines, cur = [], ""
    for w in words:
        trial = (cur + " " + w).strip()
        if draw.textlength(trial, font=fnt) <= max_width:
            cur = trial
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


def draw_bold_span(draw, pos, segments, y):
    """segments: list of (text, font, color). Draws left-to-right, returns end x."""
    px, py = pos
    for text, fnt, color in segments:
        draw.text((px, py), text, font=fnt, fill=color)
        px += draw.textlength(text, font=fnt)
    return px


def make_certificate(row, cert_id, out_path):
    name, designation, department, start_date, end_date, description, cert_date = row

    img = Image.new("RGB", (W, H), NAVY_DARK)
    draw = ImageDraw.Draw(img)
    draw_background(draw, img)

    margin = x(50)

    # header left
    draw.text((margin, x(28)), HEADER_LEFT, font=F_SMALL, fill=LIGHT_GREY)

    # QR code (top-right) — links to this certificate's verification page,
    # keyed by its permanent unique ID (not the name, so duplicate names
    # or re-issued certificates never collide)
    qr_data = f"{QR_BASE_URL}/{cert_id}"
    qr_size = x(QR_SIZE)
    qr_img = make_qr(qr_data, qr_size)
    qr_x = W - x(QR_MARGIN_RIGHT) - qr_size
    qr_y = x(QR_MARGIN_TOP)
    img.paste(qr_img, (qr_x, qr_y))

    # logo + tagline (centered)
    logo_w = draw.textlength(COMPANY_NAME, font=F_LOGO)
    draw.text(((W - logo_w) / 2, x(60)), COMPANY_NAME, font=F_LOGO, fill=WHITE)
    tag_w = draw.textlength(COMPANY_TAGLINE, font=F_TAGLINE)
    draw.text(((W - tag_w) / 2, x(112)), COMPANY_TAGLINE, font=F_TAGLINE, fill=LIGHT_GREY)

    # title
    title = "CERTIFICATE OF COMPLETION"
    title_w = draw.textlength(title, font=F_TITLE)
    draw.text(((W - title_w) / 2, x(145)), title, font=F_TITLE, fill=WHITE)

    # presented to
    presented = "This certificate is proudly presented to:"
    pw = draw.textlength(presented, font=F_PRESENTED)
    draw.text(((W - pw) / 2, x(215)), presented, font=F_PRESENTED, fill=LIGHT_GREY)

    # name + underline
    nw = draw.textlength(name, font=F_NAME)
    name_y = x(240)
    draw.text(((W - nw) / 2, name_y), name, font=F_NAME, fill=WHITE)
    line_w = max(nw + x(60), x(260))
    line_y = name_y + x(48)
    draw.line([((W - line_w) / 2, line_y), ((W + line_w) / 2, line_y)], fill=WHITE, width=2)

    # paragraph 1 (association line) — pre-text normal, designation bold, post-text normal,
    # wrapped manually word-by-word so bolding survives line breaks
    body_y = x(300)
    max_w = W - 2 * margin
    pre = "was associated with SPACEBORN as a "
    post = f" in the {department} Department from {start_date} to {end_date}."
    cy = body_y
    px = margin
    segs = [(pre, F_BODY, LIGHT_GREY), (designation, F_BODY_BOLD, WHITE), (post, F_BODY, LIGHT_GREY)]
    words_with_font = []
    for text, fnt, color in segs:
        for w in text.split(" "):
            if w:
                words_with_font.append((w, fnt, color))
    line_words = []
    line_width = 0
    space_w = draw.textlength(" ", font=F_BODY)
    for w, fnt, color in words_with_font:
        ww = draw.textlength(w, font=fnt)
        if line_width + ww + space_w > max_w and line_words:
            px = margin
            for lw, lf, lc in line_words:
                draw.text((px, cy), lw, font=lf, fill=lc)
                px += draw.textlength(lw, font=lf) + space_w
            cy += x(24)
            line_words = []
            line_width = 0
        line_words.append((w, fnt, color))
        line_width += ww + space_w
    if line_words:
        px = margin
        for lw, lf, lc in line_words:
            draw.text((px, cy), lw, font=lf, fill=lc)
            px += draw.textlength(lw, font=lf) + space_w
        cy += x(24)

    # paragraph 2 (description)
    cy += x(18)
    desc_lines = wrap_text(draw, description, F_BODY, max_w)
    for line in desc_lines:
        draw.text((margin, cy), line, font=F_BODY, fill=LIGHT_GREY)
        cy += x(22)

    # date (bottom left)
    date_y = H - x(90)
    draw.text((margin, date_y), f"Date: {cert_date}", font=F_DATE, fill=WHITE)

    # signature block (bottom right)
    sig_x = W - x(280)
    draw.text((sig_x, date_y - x(45)), ORG_LABEL, font=F_BODY_BOLD, fill=LIGHT_GREY)
    draw.text((sig_x, date_y - x(18)), "Adarsh Kumar", font=F_SIGN, fill=WHITE)
    draw.line([(sig_x, date_y + x(8)), (sig_x + x(180), date_y + x(8))], fill=ACCENT, width=1)
    draw.text((sig_x, date_y + x(14)), SIGNER_NAME, font=F_SIGNER, fill=WHITE)
    draw.text((sig_x, date_y + x(32)), SIGNER_TITLE, font=F_SIGNER_SMALL, fill=LIGHT_GREY)
    draw.text((sig_x, date_y + x(50)), SIGNER_CONTACT, font=F_SIGNER_SMALL, fill=LIGHT_GREY)

    img.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="regenerate even if PNG already exists")
    parser.add_argument("--no-sync", action="store_true", help="skip pushing certificate data to the website")
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # find (or create) the Certificate ID column
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_values = [c.value for c in header_row]
    if CERT_ID_HEADER in header_values:
        cert_id_col = header_values.index(CERT_ID_HEADER) + 1
    else:
        cert_id_col = len(header_values) + 1
        ws.cell(row=1, column=cert_id_col, value=CERT_ID_HEADER)

    made = 0
    log_rows = []
    workbook_dirty = False

    for row_cells in ws.iter_rows(min_row=2):
        values = [c.value for c in row_cells]
        if not values or not values[0]:
            continue
        vals = [str(v).strip() if v is not None else "" for v in values[:7]]
        name = vals[0]

        # get or create this row's permanent certificate ID
        id_cell = row_cells[cert_id_col - 1] if len(row_cells) >= cert_id_col else ws.cell(
            row=row_cells[0].row, column=cert_id_col
        )
        cert_id = str(id_cell.value).strip() if id_cell.value else ""
        if not cert_id:
            cert_id = new_cert_id()
            id_cell.value = cert_id
            workbook_dirty = True

        out_path = os.path.join(OUTPUT_DIR, f"{slugify(name)}-{cert_id}.png")
        if args.force or not os.path.exists(out_path):
            make_certificate(vals, cert_id, out_path)
            made += 1
            print(f"Generated: {out_path}")

        log_rows.append(
            {
                "certificate_id": cert_id,
                "name": vals[0],
                "designation": vals[1],
                "department": vals[2],
                "start_date": vals[3],
                "end_date": vals[4],
                "experience": compute_experience(vals[3], vals[4]),
                "description": vals[5],
                "cert_date": vals[6],
                "qr_url": f"{QR_BASE_URL}/{cert_id}",
                "png_filename": f"{slugify(name)}-{cert_id}.png",
            }
        )

    if workbook_dirty:
        wb.save(EXCEL_FILE)
        print(f"Assigned new certificate ID(s) — saved back to '{EXCEL_FILE}'.")

    # write/refresh the DB-ready log, covering every row every run
    if log_rows:
        with open(LOG_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(log_rows[0].keys()))
            writer.writeheader()
            writer.writerows(log_rows)
        print(f"Wrote '{LOG_FILE}' ({len(log_rows)} row(s)) for the backend team to import.")

    # push every certificate to the website so QR verify pages go live
    if log_rows and not args.no_sync:
        sync_to_website(log_rows)

    print(f"\nDone. {made} certificate(s) generated in '{OUTPUT_DIR}/'.")


if __name__ == "__main__":
    main()
